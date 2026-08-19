"""
JPXの「デリバティブ建玉残高表」から日経225オプション・日経225ミニオプションの
権利行使価格別建玉残高（当日・前日・前日比）を取得し、tidy形式のCSVで保存する。

データソース: https://www.jpx.co.jp/markets/derivatives/trading-volume/index.html
  掲載ページに当日分の *open_interest.xlsx へのリンクがあり、そのURL（コンテンツIDを
  含むパス部分）は日によって変わるため、毎回ページを読んでリンクを見つける必要がある。
  ファイル自体は当日20:00頃に更新される。土日祝・システム障害時は前営業日分が
  そのまま掲載されている場合がある。

シート構成:
  デリバティブ建玉残高状況: 限月ごとの合計（先物・金利系がメイン）
  別紙1: 日経225オプション（通常）のプット/コール、権利行使価格別
  別紙2: 日経225ミニオプションのプット/コール、権利行使価格別
  各シートの行は "NIKKEI 225 P2609-20000"（通常）/ "NK225 MINI P260820-58375"（ミニ）
  という形式の銘柄名＋続く4列（取引高, 当日建玉残高, 前日比, 前日建玉残高）。

保存先: data/oi_{YYYYMMDD}.csv （日次スナップショット、後続のダッシュボードが積み上げて使う）
"""
import csv
import datetime as dt
import os
import re
import sys
import urllib.request

UA = "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0 Safari/537.36"
INDEX_URL = "https://www.jpx.co.jp/markets/derivatives/trading-volume/index.html"
BASE = os.path.dirname(__file__)
DATA_DIR = os.path.join(BASE, "data")

CODE_RE = re.compile(r"^(NIKKEI 225|NK225 MINI)\s+([PC])(\S+)-(\d+)$")


def fetch_url(url, timeout=30):
    req = urllib.request.Request(url, headers={"User-Agent": UA})
    with urllib.request.urlopen(req, timeout=timeout) as resp:
        return resp.read()


def find_open_interest_url():
    html = fetch_url(INDEX_URL).decode("utf-8", errors="ignore")
    m = re.search(r'href="([^"]*open_interest\.xlsx)"', html)
    if not m:
        raise RuntimeError("open_interest.xlsxへのリンクが見つかりませんでした。JPXのページ構成が変わった可能性があります。")
    href = m.group(1)
    if href.startswith("http"):
        return href
    return "https://www.jpx.co.jp" + href


def parse_code(name):
    m = CODE_RE.match(name.strip())
    if not m:
        return None
    product = "standard" if m.group(1) == "NIKKEI 225" else "mini"
    put_call = "Put" if m.group(2) == "P" else "Call"
    contract = m.group(3)
    strike = int(m.group(4))
    return product, put_call, contract, strike


def parse_sheet(ws, min_row=7):
    """別紙1/別紙2形式（左ブロック=プット、右ブロック=コール、6列区切り）を読む。"""
    rows = []
    for row in ws.iter_rows(min_row=min_row, values_only=True):
        for offset in (0, 6):
            if offset >= len(row):
                continue
            name = row[offset]
            if not name or not isinstance(name, str):
                continue
            parsed = parse_code(name)
            if parsed is None:
                continue
            product, put_call, contract, strike = parsed
            volume, oi, oi_change, oi_prev = row[offset + 1], row[offset + 2], row[offset + 3], row[offset + 4]
            rows.append({
                "product": product, "put_call": put_call, "contract": contract, "strike": strike,
                "volume": volume or 0, "oi": oi or 0, "oi_change": oi_change or 0, "oi_prev": oi_prev or 0,
            })
    return rows


def main():
    import openpyxl

    date_str = sys.argv[1] if len(sys.argv) > 1 else dt.date.today().strftime("%Y%m%d")

    url = find_open_interest_url()
    print(f"取得URL: {url}")
    xlsx_bytes = fetch_url(url)

    tmp_path = os.path.join(DATA_DIR, "_tmp_open_interest.xlsx")
    os.makedirs(DATA_DIR, exist_ok=True)
    with open(tmp_path, "wb") as f:
        f.write(xlsx_bytes)

    wb = openpyxl.load_workbook(tmp_path, data_only=True)
    # ファイル自身が記載している基準日（1行目 or 2行目の日付セル）を優先して使う
    ws1 = wb["別紙1"]
    report_date = None
    for row in ws1.iter_rows(min_row=1, max_row=3, values_only=True):
        for cell in row:
            if isinstance(cell, dt.datetime):
                report_date = cell.strftime("%Y%m%d")
    if report_date is None:
        report_date = date_str

    rows = parse_sheet(wb["別紙1"]) + parse_sheet(wb["別紙2"])
    for r in rows:
        r["report_date"] = report_date

    out_path = os.path.join(DATA_DIR, f"oi_{report_date}.csv")
    fieldnames = ["report_date", "product", "put_call", "contract", "strike", "volume", "oi", "oi_change", "oi_prev"]
    with open(out_path, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=fieldnames)
        w.writeheader()
        for r in rows:
            w.writerow(r)

    os.remove(tmp_path)
    print(f"{len(rows)}行を保存: {out_path}")


if __name__ == "__main__":
    main()
